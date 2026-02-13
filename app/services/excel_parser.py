"""
Excel file parser and writer.
Reads prospect data from uploaded Excel, writes deck URLs back as output.
"""
from __future__ import annotations
import os
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook

from app.models import ProspectRow, RowResult


# Column name mappings (case-insensitive matching)
COLUMN_ALIASES = {
    "company_name": ["company name", "company", "account name", "account", "organization"],
    "industry": ["industry", "vertical", "sector"],
    "website_url": ["website url", "website", "url", "domain", "company url", "web"],
    "contact_name": ["contact name", "contact", "name", "full name", "first name"],
    "contact_title": ["contact title", "title", "job title", "role", "position"],
}


def _find_column_index(headers: list[str], field: str) -> int | None:
    """Find column index by matching against known aliases (case-insensitive)."""
    aliases = COLUMN_ALIASES.get(field, [field])
    for idx, header in enumerate(headers):
        if header and header.strip().lower() in aliases:
            return idx
    return None


def parse_excel(file_path: str) -> list[ProspectRow]:
    """
    Parse an Excel file and return a list of ProspectRow objects.

    Expects at minimum a 'Company Name' column.
    Other columns are matched flexibly via aliases.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel file must have a header row and at least one data row.")

    # Parse headers
    headers = [str(cell).strip() if cell else "" for cell in rows[0]]

    # Map columns
    col_map = {}
    for field in COLUMN_ALIASES:
        idx = _find_column_index(headers, field)
        if idx is not None:
            col_map[field] = idx

    if "company_name" not in col_map:
        raise ValueError(
            f"Could not find a 'Company Name' column. "
            f"Found headers: {headers}. "
            f"Expected one of: {COLUMN_ALIASES['company_name']}"
        )

    # Collect extra columns not mapped to known fields
    mapped_indices = set(col_map.values())
    extra_indices = [i for i in range(len(headers)) if i not in mapped_indices and headers[i]]

    prospects = []
    for row_idx, row in enumerate(rows[1:], start=2):  # 1-indexed, skip header
        company_name = row[col_map["company_name"]] if col_map.get("company_name") is not None else ""
        if not company_name or str(company_name).strip() == "":
            continue  # Skip empty rows

        # Gather extra context from unmapped columns
        extra_parts = []
        for ei in extra_indices:
            if ei < len(row) and row[ei]:
                extra_parts.append(f"{headers[ei]}: {row[ei]}")

        prospect = ProspectRow(
            row_index=row_idx,
            company_name=str(company_name).strip(),
            industry=str(row[col_map["industry"]]).strip() if col_map.get("industry") is not None and col_map["industry"] < len(row) and row[col_map["industry"]] else "",
            website_url=str(row[col_map["website_url"]]).strip() if col_map.get("website_url") is not None and col_map["website_url"] < len(row) and row[col_map["website_url"]] else "",
            contact_name=str(row[col_map["contact_name"]]).strip() if col_map.get("contact_name") is not None and col_map["contact_name"] < len(row) and row[col_map["contact_name"]] else "",
            contact_title=str(row[col_map["contact_title"]]).strip() if col_map.get("contact_title") is not None and col_map["contact_title"] < len(row) and row[col_map["contact_title"]] else "",
            extra_context="; ".join(extra_parts),
        )
        prospects.append(prospect)

    wb.close()
    return prospects


def write_results(original_file_path: str, results: list[RowResult], output_dir: str) -> str:
    """
    Copy the original Excel and add result columns (Gamma Deck URL, Status).
    Returns the path to the output file.
    """
    os.makedirs(output_dir, exist_ok=True)

    # Create output filename with timestamp
    original_name = Path(original_file_path).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"{original_name}_with_decks_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    # Copy original file
    shutil.copy2(original_file_path, output_path)

    # Open and add columns
    wb = load_workbook(output_path)
    ws = wb.active

    # Find the next empty column for our results
    max_col = ws.max_column
    url_col = max_col + 1
    status_col = max_col + 2
    pptx_col = max_col + 3

    # Write headers
    ws.cell(row=1, column=url_col, value="Gamma Deck URL")
    ws.cell(row=1, column=status_col, value="Generation Status")
    ws.cell(row=1, column=pptx_col, value="PPTX Download URL")

    # Build a map of row_index → result
    result_map = {r.row_index: r for r in results}

    for row_idx in range(2, ws.max_row + 1):
        result = result_map.get(row_idx)
        if result:
            ws.cell(row=row_idx, column=url_col, value=result.deck_url)
            ws.cell(row=row_idx, column=status_col, value=result.status.value)
            ws.cell(row=row_idx, column=pptx_col, value=result.pptx_url)

    wb.save(output_path)
    wb.close()

    return output_path
